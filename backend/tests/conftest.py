"""
Pytest fixtures and mock utilities for CRE Document Intelligence tests.
All LLM calls are mocked - no Anthropic API key consumed.
"""

import pytest
import json
import io
from unittest.mock import AsyncMock, patch, MagicMock
from typing import Any
import sys
import os

# Add backend to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))


# ============================================================================
# MOCK LLM RESPONSES
# ============================================================================

MOCK_CLASSIFICATION_RENT_ROLL = {
    "doc_type": "RENT_ROLL",
    "confidence": 0.95,
    "reasoning": "Document contains tenant names, suite numbers, square footage, and monthly rent columns typical of a rent roll."
}

MOCK_CLASSIFICATION_LEASE = {
    "doc_type": "LEASE",
    "confidence": 0.92,
    "reasoning": "Document contains lease agreement terms, tenant obligations, and landlord provisions."
}

MOCK_CLASSIFICATION_BOMA = {
    "doc_type": "BOMA",
    "confidence": 0.88,
    "reasoning": "Document contains BOMA measurement standards, RSF calculations, and load factors."
}

MOCK_CLASSIFICATION_UNKNOWN = {
    "doc_type": "UNKNOWN",
    "confidence": 0.0,
    "reasoning": "Unable to determine document type."
}

MOCK_RENT_ROLL_EXTRACTION = {
    "tenants": [
        {
            "tenant_name": "Regions Bank",
            "suite": "100",
            "rsf": 3200,
            "rent_psf": 22.50,
            "monthly_rent": 6000,
            "lease_start": "2022-01-01",
            "lease_end": "2027-12-31",
            "ar_current": 0,
            "ar_30_day": 0,
            "ar_60_day": 0,
            "ar_90_plus": 0
        },
        {
            "tenant_name": "State Farm Insurance",
            "suite": "200",
            "rsf": 2800,
            "rent_psf": 21.00,
            "monthly_rent": 4900,
            "lease_start": "2023-06-01",
            "lease_end": "2026-05-31",
            "ar_current": 4900,
            "ar_30_day": 0,
            "ar_60_day": 0,
            "ar_90_plus": 0
        },
        {
            "tenant_name": "Edward Jones",
            "suite": "300",
            "rsf": 1500,
            "rent_psf": 24.00,
            "monthly_rent": 3000,
            "lease_start": "2024-01-15",
            "lease_end": "2029-01-14",
            "ar_current": 0,
            "ar_30_day": 3000,
            "ar_60_day": 0,
            "ar_90_plus": 0
        }
    ],
    "summary": {
        "total_rsf": 7500,
        "total_monthly_rent": 13900,
        "occupancy_rate": 0.85,
        "weighted_avg_rent_psf": 22.27
    },
    "pipeline_stage": "extracted"
}

MOCK_LEASE_EXTRACTION = {
    "tenant": "Regions Bank",
    "premises": "Suite 100, 5041 Bayou Boulevard",
    "rsf": 3200,
    "term_start": "2022-01-01",
    "term_end": "2027-12-31",
    "base_rent_year_1": 72000,
    "escalation_type": "fixed",
    "escalation_rate": 0.03,
    "options": [
        {"type": "renewal", "terms": "Two 5-year options at 95% of market"}
    ],
    "cam_type": "NNN",
    "cam_cap": None,
    "ti_allowance": 15.00,
    "free_rent_months": 2,
    "guarantor": "Corporate",
    "early_termination": None,
    "pipeline_stage": "extracted"
}

MOCK_BOMA_EXTRACTION = {
    "building_totals": {
        "gross_building_area": 32000,
        "rentable_sf": 29452,
        "usable_sf": 26500,
        "common_area_factor": 1.111
    },
    "suites": [
        {"suite": "100", "usable_sf": 2880, "rentable_sf": 3200, "load_factor": 1.111},
        {"suite": "200", "usable_sf": 2520, "rentable_sf": 2800, "load_factor": 1.111},
        {"suite": "300", "usable_sf": 1350, "rentable_sf": 1500, "load_factor": 1.111}
    ],
    "measurement_standard": "BOMA 2017",
    "pipeline_stage": "extracted"
}

MOCK_SYNTHESIS = {
    "deal_score": {
        "overall": 66,
        "deal_readiness": "YELLOW",
        "components": {
            "documentation": 70,
            "data_quality": 65,
            "risk": 63
        }
    },
    "rsf_recovery": {
        "boma_rsf": 29452,
        "rent_roll_rsf": 24847,
        "delta_sf": 4605,
        "delta_pct": 18.53,
        "annual_recovery_potential": 51162
    },
    "lease_audit": {
        "walt_months": 22.4,
        "near_term_rollover_pct": 0.15,
        "below_market_leases": 1
    },
    "tenant_summary": {
        "total_tenants": 3,
        "occupied_sf": 7500,
        "vacancy_sf": 2000,
        "occupancy_pct": 0.79
    },
    "what_to_get_next": [
        "Executed lease for State Farm (expiring in 12 months)",
        "Updated BOMA measurement report",
        "Historical operating statements (3 years)"
    ]
}

MOCK_ARITHMETIC_VERIFICATION = {
    "checks": [
        {
            "check_type": "rsf_total",
            "expected": 7500,
            "actual": 7500,
            "status": "VERIFIED",
            "variance_pct": 0
        },
        {
            "check_type": "rent_psf",
            "tenant": "Regions Bank",
            "expected": 22.50,
            "actual": 22.50,
            "status": "VERIFIED"
        },
        {
            "check_type": "boma_vs_rent_roll",
            "boma_rsf": 29452,
            "rent_roll_rsf": 24847,
            "delta_pct": 18.53,
            "status": "CALC_MISMATCH",
            "note": "RSF variance exceeds 5% threshold"
        }
    ],
    "total": 3,
    "mismatches": 1,
    "verified": 2
}

MOCK_RED_FLAGS = {
    "red_flags": [
        {
            "severity": "CRITICAL",
            "category": "rsf_discrepancy",
            "title": "RSF Variance Exceeds 15%",
            "description": "BOMA RSF (29,452) differs from Rent Roll RSF (24,847) by 18.53%",
            "recommendation": "Request updated BOMA measurement or reconcile with landlord"
        },
        {
            "severity": "CRITICAL",
            "category": "near_term_rollover",
            "title": "Major Tenant Expiration Within 12 Months",
            "description": "State Farm lease expires 2026-05-31 representing 15% of rent",
            "recommendation": "Obtain renewal status or replacement tenant pipeline"
        },
        {
            "severity": "HIGH",
            "category": "ar_aging",
            "title": "30+ Day Receivables",
            "description": "Edward Jones has $3,000 in 30-day AR",
            "recommendation": "Review collection status and tenant creditworthiness"
        },
        {
            "severity": "MEDIUM",
            "category": "missing_documents",
            "title": "Historical Operating Statements Missing",
            "description": "No T-12 or historical financials provided",
            "recommendation": "Request 3 years of operating statements"
        }
    ],
    "summary": {
        "critical": 2,
        "high": 1,
        "medium": 1,
        "low": 0
    }
}

MOCK_OCR_CLEANED_TEXT = """RENT ROLL - Town & Country Plaza
As of February 2026

Suite    Tenant Name           RSF      Rent/SF    Monthly Rent
100      Regions Bank         3,200     $22.50        $6,000
200      State Farm           2,800     $21.00        $4,900
300      Edward Jones         1,500     $24.00        $3,000

Total Occupied:  7,500 SF
Vacancy:         2,000 SF
Total Building: 9,500 SF
"""


def mock_llm_router(prompt: str, **kwargs) -> str:
    """
    Routes mock responses based on prompt content.
    This is the core mock that replaces all LLM calls.
    """
    prompt_lower = prompt.lower()
    
    # Classification prompts
    if "classify" in prompt_lower or "document type" in prompt_lower:
        if "rent roll" in prompt_lower or "tenant" in prompt_lower and "suite" in prompt_lower:
            return json.dumps(MOCK_CLASSIFICATION_RENT_ROLL)
        elif "lease" in prompt_lower and ("agreement" in prompt_lower or "term" in prompt_lower):
            return json.dumps(MOCK_CLASSIFICATION_LEASE)
        elif "boma" in prompt_lower or "measurement" in prompt_lower:
            return json.dumps(MOCK_CLASSIFICATION_BOMA)
        else:
            return json.dumps(MOCK_CLASSIFICATION_RENT_ROLL)  # Default for tests
    
    # Extraction prompts
    if "extract" in prompt_lower:
        if "rent roll" in prompt_lower or "rent_roll" in prompt_lower:
            return json.dumps(MOCK_RENT_ROLL_EXTRACTION)
        elif "lease" in prompt_lower:
            return json.dumps(MOCK_LEASE_EXTRACTION)
        elif "boma" in prompt_lower:
            return json.dumps(MOCK_BOMA_EXTRACTION)
        else:
            return json.dumps(MOCK_RENT_ROLL_EXTRACTION)
    
    # Synthesis prompts
    if "synthesize" in prompt_lower or "synthesis" in prompt_lower or "combine" in prompt_lower:
        return json.dumps(MOCK_SYNTHESIS)
    
    # Arithmetic verification
    if "arithmetic" in prompt_lower or "verify" in prompt_lower or "calculation" in prompt_lower:
        return json.dumps(MOCK_ARITHMETIC_VERIFICATION)
    
    # Red flags
    if "red flag" in prompt_lower or "risk" in prompt_lower:
        return json.dumps(MOCK_RED_FLAGS)
    
    # OCR cleaning
    if "ocr" in prompt_lower or "clean" in prompt_lower:
        return MOCK_OCR_CLEANED_TEXT
    
    # Default - return synthesis
    return json.dumps(MOCK_SYNTHESIS)


async def async_mock_llm_router(prompt: str, **kwargs) -> str:
    """Async version of the mock router."""
    return mock_llm_router(prompt, **kwargs)


# ============================================================================
# FIXTURES
# ============================================================================

@pytest.fixture
def mock_llm():
    """Fixture that patches BaseAgent.call_llm with mock responses."""
    with patch('agents.base.BaseAgent.call_llm', AsyncMock(side_effect=async_mock_llm_router)):
        yield


@pytest.fixture
def sample_rent_roll_text():
    """Sample rent roll document text."""
    return """
    RENT ROLL
    Property: 5041 Bayou Boulevard
    As of: March 2026
    
    Suite   Tenant Name          RSF      Rent/SF   Monthly Rent   Lease Start   Lease End
    100     Regions Bank        3,200     $22.50      $6,000       01/01/2022    12/31/2027
    200     State Farm Ins      2,800     $21.00      $4,900       06/01/2023    05/31/2026
    300     Edward Jones        1,500     $24.00      $3,000       01/15/2024    01/14/2029
    
    TOTALS                      7,500                $13,900
    
    Vacancy: Suite 400 - 2,000 SF
    """


@pytest.fixture
def sample_lease_text():
    """Sample lease document text."""
    return """
    COMMERCIAL LEASE AGREEMENT
    
    LANDLORD: Bayou Properties LLC
    TENANT: Regions Bank
    PREMISES: Suite 100, 5041 Bayou Boulevard, Pensacola, FL 32503
    
    1. TERM: January 1, 2022 through December 31, 2027 (6 years)
    2. RENTABLE SQUARE FEET: 3,200 RSF
    3. BASE RENT: $72,000 per year ($6,000/month, $22.50/SF)
    4. ESCALATION: 3% annual increases
    5. LEASE TYPE: NNN (Triple Net)
    6. TENANT IMPROVEMENTS: $15.00/SF allowance
    7. FREE RENT: 2 months at commencement
    8. RENEWAL OPTIONS: Two (2) five-year options at 95% of then-market rent
    9. GUARANTOR: Corporate guarantee
    """


@pytest.fixture
def sample_boma_text():
    """Sample BOMA measurement document text."""
    return """
    BOMA MEASUREMENT REPORT
    Property: 5041 Bayou Boulevard
    Standard: BOMA 2017 Office
    
    BUILDING SUMMARY
    Gross Building Area:     32,000 SF
    Rentable Square Feet:    29,452 SF
    Usable Square Feet:      26,500 SF
    Common Area Factor:      1.111
    
    SUITE MEASUREMENTS
    Suite 100:  Usable 2,880 SF  |  Rentable 3,200 SF  |  Load Factor 1.111
    Suite 200:  Usable 2,520 SF  |  Rentable 2,800 SF  |  Load Factor 1.111
    Suite 300:  Usable 1,350 SF  |  Rentable 1,500 SF  |  Load Factor 1.111
    Suite 400:  Usable 1,800 SF  |  Rentable 2,000 SF  |  Load Factor 1.111 (VACANT)
    """


@pytest.fixture
def sample_pdf_bytes():
    """Simulated PDF file bytes (text-based, not scanned)."""
    # This is a minimal PDF structure with embedded text
    return b"""%PDF-1.4
1 0 obj
<< /Type /Catalog /Pages 2 0 R >>
endobj
2 0 obj
<< /Type /Pages /Kids [3 0 R] /Count 1 >>
endobj
3 0 obj
<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Contents 4 0 R >>
endobj
4 0 obj
<< /Length 44 >>
stream
BT /F1 12 Tf 100 700 Td (RENT ROLL) Tj ET
endstream
endobj
xref
0 5
0000000000 65535 f
0000000009 00000 n
0000000058 00000 n
0000000115 00000 n
0000000206 00000 n
trailer
<< /Size 5 /Root 1 0 R >>
startxref
300
%%EOF
"""


@pytest.fixture
def sample_excel_bytes():
    """Create a real Excel file for testing."""
    try:
        import openpyxl
        from io import BytesIO
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rent Roll"
        
        # Headers
        headers = ["Suite", "Tenant", "RSF", "Rent/SF", "Monthly Rent"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        data = [
            ["100", "Regions Bank", 3200, 22.50, 6000],
            ["200", "State Farm", 2800, 21.00, 4900],
            ["300", "Edward Jones", 1500, 24.00, 3000],
        ]
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    except ImportError:
        # Return minimal xlsx bytes if openpyxl not available
        return b"PK\x03\x04"  # ZIP signature (xlsx is a zip)


@pytest.fixture
def sample_csv_bytes():
    """Sample CSV file bytes."""
    return b"""Suite,Tenant,RSF,Rent/SF,Monthly Rent
100,Regions Bank,3200,22.50,6000
200,State Farm,2800,21.00,4900
300,Edward Jones,1500,24.00,3000
"""


@pytest.fixture
def initial_pipeline_state():
    """Initial state for LangGraph pipeline tests."""
    return {
        "deal_id": "test-deal-001",
        "deal_name": "5041 Bayou Boulevard",
        "raw_files": {},
        "file_content_types": {},
        "raw_documents": [],
        "ingest_errors": [],
        "classified_documents": [],
        "classification_errors": [],
        "extractions": [],
        "extraction_errors": [],
        "synthesis": {},
        "rsf_recovery": {},
        "score_summary": {},
        "synthesis_error": None,
        "arithmetic_verification": {},
        "rent_roll_analysis": {},
        "rsf_reconciliation": {},
        "red_flags_result": {},
        "deal_score_result": {},
        "completeness_result": {},
        "pipeline_stage": "pending",
        "pipeline_errors": [],
        "completed_at": None,
    }


@pytest.fixture
def test_client():
    """Create test client for API endpoint tests."""
    try:
        from httpx import AsyncClient, ASGITransport
        from main import app
        
        transport = ASGITransport(app=app)
        return AsyncClient(transport=transport, base_url="http://test")
    except ImportError:
        pytest.skip("httpx not installed")
