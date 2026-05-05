"""
Document Parsing Agent
Handles document ingestion, text extraction, and OCR detection.
"""

import io
from dataclasses import dataclass
from typing import Optional

try:
    import PyPDF2
except ImportError:
    PyPDF2 = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


@dataclass
class ParseResult:
    """Result of document parsing check."""
    is_parseable: bool
    file_type: str
    extracted_text: str
    page_count: Optional[int]
    needs_ocr: bool
    method: str
    reason: str


class DocumentParsingAgent:
    """Agent for document parsing and text extraction."""
    
    def __init__(self):
        self.name = "DocumentParsingAgent"
    
    def check(
        self,
        filename: str,
        file_content: bytes,
        content_type: str
    ) -> ParseResult:
        """
        Check if a document can be parsed and extract text if possible.
        Returns ParseResult with extracted text or OCR requirement.
        """
        filename_lower = filename.lower()
        
        # PDF handling
        if filename_lower.endswith('.pdf') or content_type == 'application/pdf':
            return self._parse_pdf(file_content, filename)
        
        # Excel handling
        if filename_lower.endswith(('.xlsx', '.xls')) or 'spreadsheet' in content_type or 'excel' in content_type:
            return self._parse_excel(file_content, filename)
        
        # CSV handling
        if filename_lower.endswith('.csv') or content_type == 'text/csv':
            return self._parse_csv(file_content, filename)
        
        # Text handling
        if filename_lower.endswith('.txt') or content_type.startswith('text/'):
            return self._parse_text(file_content, filename)
        
        # Image handling - needs OCR
        if filename_lower.endswith(('.png', '.jpg', '.jpeg', '.tiff', '.bmp')) or content_type.startswith('image/'):
            return ParseResult(
                is_parseable=True,
                file_type="image",
                extracted_text="",
                page_count=1,
                needs_ocr=True,
                method="ocr_required",
                reason="Image file requires OCR processing"
            )
        
        return ParseResult(
            is_parseable=False,
            file_type="unknown",
            extracted_text="",
            page_count=None,
            needs_ocr=False,
            method="none",
            reason=f"Unsupported file type: {content_type}"
        )
    
    def _parse_pdf(self, file_content: bytes, filename: str) -> ParseResult:
        """Extract text from PDF."""
        if PyPDF2 is None:
            return ParseResult(
                is_parseable=True,
                file_type="pdf",
                extracted_text="",
                page_count=None,
                needs_ocr=True,
                method="ocr_fallback",
                reason="PyPDF2 not available, falling back to OCR"
            )
        
        try:
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
            page_count = len(pdf_reader.pages)
            
            text_parts = []
            for page in pdf_reader.pages:
                text = page.extract_text() or ""
                text_parts.append(text)
            
            extracted_text = "\n\n".join(text_parts)
            
            # Check if we got meaningful text or if OCR is needed
            if len(extracted_text.strip()) < 100:
                return ParseResult(
                    is_parseable=True,
                    file_type="pdf",
                    extracted_text="",
                    page_count=page_count,
                    needs_ocr=True,
                    method="ocr_required",
                    reason="PDF appears to be scanned/image-based, OCR required"
                )
            
            return ParseResult(
                is_parseable=True,
                file_type="pdf",
                extracted_text=extracted_text,
                page_count=page_count,
                needs_ocr=False,
                method="pypdf2",
                reason="Text extracted successfully"
            )
            
        except Exception as e:
            return ParseResult(
                is_parseable=True,
                file_type="pdf",
                extracted_text="",
                page_count=None,
                needs_ocr=True,
                method="ocr_fallback",
                reason=f"PDF parsing failed: {str(e)}, falling back to OCR"
            )
    
    def _parse_excel(self, file_content: bytes, filename: str) -> ParseResult:
        """Extract text from Excel files."""
        if openpyxl is None:
            return ParseResult(
                is_parseable=False,
                file_type="excel",
                extracted_text="",
                page_count=None,
                needs_ocr=False,
                method="none",
                reason="openpyxl not available for Excel parsing"
            )
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            
            text_parts = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"=== Sheet: {sheet_name} ===")
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        text_parts.append(row_text)
            
            extracted_text = "\n".join(text_parts)
            
            return ParseResult(
                is_parseable=True,
                file_type="excel",
                extracted_text=extracted_text,
                page_count=len(workbook.sheetnames),
                needs_ocr=False,
                method="openpyxl",
                reason="Excel data extracted successfully"
            )
            
        except Exception as e:
            return ParseResult(
                is_parseable=False,
                file_type="excel",
                extracted_text="",
                page_count=None,
                needs_ocr=False,
                method="none",
                reason=f"Excel parsing failed: {str(e)}"
            )
    
    def _parse_csv(self, file_content: bytes, filename: str) -> ParseResult:
        """Extract text from CSV files."""
        try:
            text = file_content.decode('utf-8', errors='replace')
            return ParseResult(
                is_parseable=True,
                file_type="csv",
                extracted_text=text,
                page_count=1,
                needs_ocr=False,
                method="decode",
                reason="CSV text extracted successfully"
            )
        except Exception as e:
            return ParseResult(
                is_parseable=False,
                file_type="csv",
                extracted_text="",
                page_count=None,
                needs_ocr=False,
                method="none",
                reason=f"CSV parsing failed: {str(e)}"
            )
    
    def _parse_text(self, file_content: bytes, filename: str) -> ParseResult:
        """Extract text from plain text files."""
        try:
            text = file_content.decode('utf-8', errors='replace')
            return ParseResult(
                is_parseable=True,
                file_type="text",
                extracted_text=text,
                page_count=1,
                needs_ocr=False,
                method="decode",
                reason="Text extracted successfully"
            )
        except Exception as e:
            return ParseResult(
                is_parseable=False,
                file_type="text",
                extracted_text="",
                page_count=None,
                needs_ocr=False,
                method="none",
                reason=f"Text parsing failed: {str(e)}"
            )
