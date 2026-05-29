import io
from typing import Optional
from datetime import datetime
import uuid

# PDF and Excel processing
try:
    from PyPDF2 import PdfReader
except ImportError:
    PdfReader = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import pandas as pd
except ImportError:
    pd = None

from agents import (
    DocumentClassifierAgent,
    LeaseAbstractionAgent,
    RentRollAgent,
    RSFReconciliationAgent,
    RiskScoringAgent,
    RedFlagDetectionAgent,
)
from models.schemas import (
    DocumentType,
    ProcessingStatus,
    ProcessedDocument,
    AnalysisResult,
)


class DocumentProcessor:
    """Orchestrates document processing and AI agent analysis."""
    
    def __init__(self):
        self.classifier = DocumentClassifierAgent()
        self.lease_agent = LeaseAbstractionAgent()
        self.rent_roll_agent = RentRollAgent()
        self.rsf_agent = RSFReconciliationAgent()
        self.risk_agent = RiskScoringAgent()
        self.red_flag_agent = RedFlagDetectionAgent()
        
        # In-memory storage (use database in production)
        self.documents: dict[str, ProcessedDocument] = {}
        self.deals: dict[str, dict] = {}
    
    def extract_text_from_pdf(self, file_content: bytes) -> str:
        """Extract text from PDF file."""
        if PdfReader is None:
            return "[PDF parsing not available - PyPDF2 not installed]"
        
        try:
            pdf_file = io.BytesIO(file_content)
            reader = PdfReader(pdf_file)
            text = ""
            for page in reader.pages:
                text += page.extract_text() or ""
            return text
        except Exception as e:
            return f"[Error extracting PDF text: {str(e)}]"
    
    def extract_data_from_excel(self, file_content: bytes) -> str:
        """Extract data from Excel file as text representation."""
        if openpyxl is None or pd is None:
            return "[Excel parsing not available - openpyxl/pandas not installed]"
        
        try:
            excel_file = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            
            all_sheets_text = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = f"=== Sheet: {sheet_name} ===\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
                    sheet_text += row_text + "\n"
                
                all_sheets_text.append(sheet_text)
            
            return "\n\n".join(all_sheets_text)
        except Exception as e:
            return f"[Error extracting Excel data: {str(e)}]"
    
    async def process_document(
        self, 
        filename: str, 
        file_content: bytes,
        content_type: str
    ) -> ProcessedDocument:
        """Process an uploaded document."""
        doc_id = str(uuid.uuid4())
        
        # Extract text based on file type
        if content_type == "application/pdf" or filename.endswith(".pdf"):
            extracted_text = self.extract_text_from_pdf(file_content)
            page_count = len(PdfReader(io.BytesIO(file_content)).pages) if PdfReader else None
        elif content_type in [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel"
        ] or filename.endswith((".xlsx", ".xls")):
            extracted_text = self.extract_data_from_excel(file_content)
            page_count = None
        else:
            extracted_text = file_content.decode("utf-8", errors="ignore")
            page_count = None
        
        # Classify document
        classification = await self.classifier.classify_document(extracted_text, filename)
        
        doc_type = DocumentType.OTHER
        if "classification" in classification:
            type_str = classification["classification"].get("document_type", "OTHER")
            try:
                doc_type = DocumentType(type_str.lower())
            except ValueError:
                doc_type = DocumentType.OTHER
        
        # Create document record
        doc = ProcessedDocument(
            id=doc_id,
            filename=filename,
            document_type=doc_type,
            status=ProcessingStatus.COMPLETED,
            uploaded_at=datetime.utcnow().isoformat(),
            processed_at=datetime.utcnow().isoformat(),
            page_count=page_count,
            extracted_data={
                "text_preview": extracted_text[:1000],
                "classification": classification,
                "full_text": extracted_text,
            }
        )
        
        self.documents[doc_id] = doc
        return doc
    
    async def run_full_analysis(self, deal_id: str, document_ids: list[str]) -> AnalysisResult:
        """Run full analysis pipeline on a set of documents."""
        
        # Gather documents
        docs = [self.documents[doc_id] for doc_id in document_ids if doc_id in self.documents]
        
        # Separate by type
        leases = [d for d in docs if d.document_type == DocumentType.LEASE]
        rent_rolls = [d for d in docs if d.document_type == DocumentType.RENT_ROLL]
        boma_reports = [d for d in docs if d.document_type == DocumentType.BOMA_MEASUREMENT]
        
        # Extract lease abstracts
        lease_abstracts = []
        for lease_doc in leases:
            if lease_doc.extracted_data and "full_text" in lease_doc.extracted_data:
                result = await self.lease_agent.extract_lease(
                    lease_doc.extracted_data["full_text"]
                )
                if "lease_abstract" in result:
                    lease_abstracts.append(result["lease_abstract"])
        
        # Parse rent roll
        rent_roll_analysis = {}
        for rr_doc in rent_rolls:
            if rr_doc.extracted_data and "full_text" in rr_doc.extracted_data:
                rent_roll_analysis = await self.rent_roll_agent.parse_rent_roll(
                    rr_doc.extracted_data["full_text"]
                )
                break  # Use first rent roll
        
        # RSF reconciliation
        boma_data = None
        for boma_doc in boma_reports:
            if boma_doc.extracted_data and "full_text" in boma_doc.extracted_data:
                boma_data = {"text": boma_doc.extracted_data["full_text"]}
                break
        
        rsf_reconciliation = await self.rsf_agent.reconcile(
            rent_roll_data=rent_roll_analysis,
            lease_data=lease_abstracts,
            boma_data=boma_data
        )
        
        # Red flag detection
        red_flags_result = await self.red_flag_agent.detect_flags(
            rent_roll_analysis=rent_roll_analysis,
            lease_abstracts=lease_abstracts,
            rsf_reconciliation=rsf_reconciliation
        )
        
        # Document completeness
        completeness = await self.classifier.assess_completeness(
            [{"type": d.document_type.value, "filename": d.filename} for d in docs]
        )
        
        # Risk scoring
        deal_score = await self.risk_agent.score_deal(
            documents_status=completeness,
            rsf_reconciliation=rsf_reconciliation,
            lease_abstracts=lease_abstracts,
            rent_roll_analysis=rent_roll_analysis,
            red_flags=red_flags_result.get("red_flags", [])
        )
        
        # Build result
        from models.schemas import DealScore, RSFReconciliation, RedFlag
        
        result = AnalysisResult(
            deal_id=deal_id,
            property_name="Property Under Analysis",
            property_address="Address TBD",
            analysis_date=datetime.utcnow().isoformat(),
            deal_score=DealScore(
                overall_score=deal_score.get("deal_score", {}).get("overall_score", 75),
                tier=deal_score.get("deal_score", {}).get("tier", "Standard"),
                sub_scores=deal_score.get("deal_score", {}).get("sub_scores", {}),
                score_factors=deal_score.get("score_factors", [])
            ),
            rsf_reconciliation=RSFReconciliation(
                total_rsf_rent_roll=rsf_reconciliation.get("reconciliation", {}).get("total_rsf_rent_roll", 0),
                total_rsf_leases=rsf_reconciliation.get("reconciliation", {}).get("total_rsf_leases", 0),
                total_rsf_boma=rsf_reconciliation.get("reconciliation", {}).get("total_rsf_boma", 0),
                variance_rent_roll_vs_boma=rsf_reconciliation.get("reconciliation", {}).get("variance_rent_roll_vs_boma", 0),
                variance_percentage=rsf_reconciliation.get("reconciliation", {}).get("variance_percentage", 0),
                estimated_annual_revenue_impact=rsf_reconciliation.get("reconciliation", {}).get("estimated_annual_revenue_impact", 0),
                discrepancies=rsf_reconciliation.get("by_tenant", [])
            ),
            tenants=[],  # Would be populated from rent roll
            lease_abstracts=[],  # Would be converted from lease_abstracts
            red_flags=[],  # Would be converted from red_flags_result
            documents_processed=docs,
            what_to_get_next=completeness.get("what_to_request_next", []),
            financial_summary=rent_roll_analysis.get("summary", {})
        )
        
        self.deals[deal_id] = {"result": result, "raw_data": {
            "lease_abstracts": lease_abstracts,
            "rent_roll": rent_roll_analysis,
            "rsf": rsf_reconciliation,
            "red_flags": red_flags_result,
            "score": deal_score
        }}
        
        return result


# Singleton instance
processor = DocumentProcessor()
